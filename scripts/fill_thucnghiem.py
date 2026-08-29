"""Fill the per-image system results into ThucNghiem.xlsx.

Reads the per-image evaluation CSV (default ``results/eval_baseline.csv``) and
writes the system's top-1 correctness (1 = correct, 0 = wrong) into the two
"HỆ THỐNG" columns of ThucNghiem.xlsx, keyed by prompt_id (STT):

    images in data/ChatGPT/ → column "HỆ THỐNG CHAT"
    images in data/Gemini/  → column "HỆ THỐNG GEMINI"

Only those two columns are touched; every other cell is left as-is. The target
columns are located by matching their header text in the sub-header row, so the
script keeps working if the sheet layout shifts.

Usage (from project root, with ThucNghiem.xlsx CLOSED in Excel):
    python scripts/fill_thucnghiem.py
    python scripts/fill_thucnghiem.py --csv results/eval_baseline.csv --xlsx ThucNghiem.xlsx
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_CHAT_HEADER = "HỆ THỐNG CHAT"
_GEMINI_HEADER = "HỆ THỐNG GEMINI"


def _load_top1(csv_path: Path) -> Dict[int, Dict[str, int]]:
    """Map prompt_id → {"chatgpt": 0/1, "gemini": 0/1} from the eval CSV.

    Rows with an error or no candidates (API/quota failures) are skipped — they
    were never really analysed, so no 0/1 is written for them.
    """
    result: Dict[int, Dict[str, int]] = {}
    with csv_path.open(encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            if row.get("error") or int(row.get("n_candidates") or 0) <= 0:
                continue
            try:
                pid = int(row["prompt_id"])
            except (KeyError, ValueError):
                continue
            gen = row["generator"].strip().lower()
            result.setdefault(pid, {})[gen] = int(row["top1_hit"])
    return result


def _find_columns(ws) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """Locate the STT column and the two target columns by header text.

    Returns ``(stt_col, chat_col, gemini_col)`` as 1-based column indices, or
    None for any not found. Scans the first ~5 rows for the header cells.
    """
    stt_col = chat_col = gemini_col = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if not isinstance(val, str):
                continue
            text = val.strip()
            if text == _CHAT_HEADER:
                chat_col = c
            elif text == _GEMINI_HEADER:
                gemini_col = c
            elif text.upper() == "STT":
                stt_col = c
    return stt_col, chat_col, gemini_col


def main() -> int:
    """Fill the two system-result columns from the eval CSV."""
    parser = argparse.ArgumentParser(description="Fill ThucNghiem.xlsx 1/0 results")
    parser.add_argument("--csv", default=str(_PROJECT_ROOT / "results" / "eval_baseline.csv"))
    parser.add_argument("--xlsx", default=str(_PROJECT_ROOT / "ThucNghiem.xlsx"))
    args = parser.parse_args()

    csv_path, xlsx_path = Path(args.csv), Path(args.xlsx)
    if not csv_path.is_file():
        print(f"ERROR: CSV not found: {csv_path}")
        return 1
    if not xlsx_path.is_file():
        print(f"ERROR: workbook not found: {xlsx_path}")
        return 1

    top1 = _load_top1(csv_path)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    stt_col, chat_col, gemini_col = _find_columns(ws)
    if not (stt_col and chat_col and gemini_col):
        print(f"ERROR: could not locate columns "
              f"(STT={stt_col}, '{_CHAT_HEADER}'={chat_col}, "
              f"'{_GEMINI_HEADER}'={gemini_col})")
        return 1

    written = 0
    for r in range(1, ws.max_row + 1):
        stt = ws.cell(row=r, column=stt_col).value
        if not isinstance(stt, int) or stt not in top1:
            continue
        scores = top1[stt]
        if "chatgpt" in scores:
            ws.cell(row=r, column=chat_col).value = scores["chatgpt"]
        if "gemini" in scores:
            ws.cell(row=r, column=gemini_col).value = scores["gemini"]
        written += 1

    try:
        wb.save(xlsx_path)
    except PermissionError:
        print(f"ERROR: cannot write {xlsx_path} — close it in Excel first.")
        return 1
    chat_sum = sum(s.get("chatgpt", 0) for s in top1.values())
    gem_sum = sum(s.get("gemini", 0) for s in top1.values())
    print(f"Filled {written} rows in {xlsx_path.name} "
          f"(cols: CHAT=#{chat_col}, GEMINI=#{gemini_col}). "
          f"Correct: HỆ THỐNG CHAT={chat_sum}/{len(top1)}, "
          f"HỆ THỐNG GEMINI={gem_sum}/{len(top1)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
