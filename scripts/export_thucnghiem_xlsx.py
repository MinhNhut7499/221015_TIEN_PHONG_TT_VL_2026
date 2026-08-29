"""Export architecture-style evaluation results into researcher-friendly .xlsx.

Reads the per-image evaluation CSV(s) produced by ``evaluate_openvocab.py`` and
writes clean, easy-to-read Excel workbooks — the way experimental results are
usually reported in a paper.

Two modes:

    run        one eval CSV  -> one workbook with 2 sheets:
                 "Tong hop"        : top-1 accuracy (overall + by image source)
                                     and a risk-coverage sweep.
                 "Per-image detail" : one row per image, colour-coded Correct/Wrong.

    aggregate  several eval CSVs -> one workbook comparing the runs:
                 one row per run + a final "mean ± standard deviation" row,
                 plus a risk-coverage table averaged over the runs.

Metrics are recomputed from the CSV with exactly the same rule the harness uses
(rows with an API error or zero candidates are excluded from scoring), so the
numbers match ``evaluate_openvocab.py`` and ``generate_baocao_methodology.py``.

Usage (from project root, with the target .xlsx CLOSED in Excel):
    python scripts/export_thucnghiem_xlsx.py --mode run \
        --csv results/eval_run2.csv --xlsx results/thucnghiem_5runs/KetQua_Lan2.xlsx \
        --label "Run 2 (2026-07-05)"

    python scripts/export_thucnghiem_xlsx.py --mode aggregate \
        --csvs results/eval_thucnghiem_full.csv,results/eval_run2.csv \
        --labels "Run 1 (2026-06-21);Run 2" \
        --xlsx results/thucnghiem_5runs/TongHop_5Lan.xlsx
"""
from __future__ import annotations

import argparse
import csv
import statistics
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
_RISK_THRESHOLDS = [0.0, 0.2, 0.4, 0.5, 0.6, 0.7, 0.8]

# ---- shared styles ---------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
_SUB_FONT = Font(italic=True, size=10, color="595959")
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_TOTAL_FONT = Font(bold=True)
_OK_FILL = PatternFill("solid", fgColor="C6EFCE")     # green
_OK_FONT = Font(color="006100")
_BAD_FILL = PatternFill("solid", fgColor="FFC7CE")    # red
_BAD_FONT = Font(color="9C0006")
_ERR_FILL = PatternFill("solid", fgColor="D9D9D9")    # grey
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


@dataclass
class RunMetrics:
    """Top-1 focused aggregate metrics recomputed from one eval CSV."""

    n_scored: int
    n_errors: int
    top1: int
    by_gen: Dict[str, Dict[str, int]] = field(default_factory=dict)
    risk_rows: List[Tuple[float, int, int]] = field(default_factory=list)
    detail: List[dict] = field(default_factory=list)  # all raw rows (incl. errors)

    def pct(self, num: int, den: Optional[int] = None) -> float:
        """Return percentage num/den; den defaults to the scored image count."""
        d = self.n_scored if den is None else den
        return 100.0 * num / d if d else 0.0


def _load_metrics(csv_path: Path) -> RunMetrics:
    """Recompute top-1 metrics from a per-image results CSV (harness rule)."""
    with csv_path.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    # An API/quota failure produced no analysis; counting it as a wrong answer
    # would pollute the accuracy, so it is excluded from scoring.
    scored = [
        r for r in rows
        if not r.get("error") and int(r.get("n_candidates") or 0) > 0
    ]

    by_gen: Dict[str, Dict[str, int]] = {}
    for r in scored:
        g = by_gen.setdefault(r["generator"].strip().lower(), {"n": 0, "top1": 0})
        g["n"] += 1
        g["top1"] += int(r["top1_hit"])

    have_agree = [
        r for r in scored if r.get("panel_agreement") not in (None, "", "None")
    ]
    risk_rows: List[Tuple[float, int, int]] = []
    for t in _RISK_THRESHOLDS:
        kept = [r for r in have_agree if float(r["panel_agreement"]) >= t]
        sel_top1 = sum(int(r["top1_hit"]) for r in kept)
        risk_rows.append((t, len(kept), sel_top1))

    return RunMetrics(
        n_scored=len(scored),
        n_errors=len(rows) - len(scored),
        top1=sum(int(r["top1_hit"]) for r in scored),
        by_gen=by_gen,
        risk_rows=risk_rows,
        detail=rows,
    )


# ---- small worksheet helpers ----------------------------------------------
def _autofit(ws: Worksheet, widths: Dict[int, int]) -> None:
    """Set fixed column widths (1-based col index -> width)."""
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _header_row(ws: Worksheet, row: int, headers: List[str]) -> None:
    """Write a styled header row starting at column 1."""
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _pct_str(num: int, den: int) -> str:
    """Format 'num/den (xx.x%)'."""
    return f"{num}/{den} ({100.0 * num / den:.1f}%)" if den else "0/0 (n/a)"


# ---- sheet 1: summary ------------------------------------------------------
def _write_summary_sheet(ws: Worksheet, m: RunMetrics, label: str) -> None:
    """Write the top-1 + risk-coverage summary sheet."""
    ws.title = "Summary"
    ws["A1"] = f"EXPERIMENT RESULTS - {label}"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = (f"Scored images: {m.n_scored}   |   "
                f"API errors (excluded from scoring): {m.n_errors}")
    ws["A2"].font = _SUB_FONT

    # Block 1: top-1 accuracy overall + by source.
    ws["A4"] = "Top-1 accuracy (correct style ranked #1)"
    ws["A4"].font = Font(bold=True, size=12)
    _header_row(ws, 5, ["Group", "Correct", "Scored", "Top-1 (%)"])

    def _acc_row(r: int, name: str, top1: int, n: int, highlight: bool) -> None:
        vals = [name, top1, n, round(100.0 * top1 / n, 1) if n else 0.0]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _LEFT if col == 1 else _CENTER
            if highlight:
                cell.fill = _TOTAL_FILL
                cell.font = _TOTAL_FONT

    row = 6
    _acc_row(row, "OVERALL (ChatGPT + Gemini)", m.top1, m.n_scored, True)
    row += 1
    for gen in ("chatgpt", "gemini"):
        g = m.by_gen.get(gen)
        if g:
            nice = "ChatGPT images" if gen == "chatgpt" else "Gemini images"
            _acc_row(row, nice, g["top1"], g["n"], False)
            row += 1

    # Block 2: risk-coverage sweep.
    row += 1
    ws.cell(row=row, column=1,
            value="Risk-coverage (abstain when panel agreement < threshold t)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    _header_row(ws, row, ["Threshold t", "Kept",
                          "Coverage (%)", "Selective Top-1 (%)"])
    row += 1
    for t, kept, sel in m.risk_rows:
        cov = round(100.0 * kept / m.n_scored, 1) if m.n_scored else 0.0
        sel_pct = round(100.0 * sel / kept, 1) if kept else 0.0
        for col, v in enumerate([t, kept, cov, sel_pct], start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _CENTER
        row += 1

    _autofit(ws, {1: 30, 2: 16, 3: 16, 4: 18})
    ws.sheet_view.showGridLines = False


# ---- sheet 2: per-image detail --------------------------------------------
def _write_detail_sheet(ws: Worksheet, m: RunMetrics) -> None:
    """Write one colour-coded row per image (Correct / Wrong / API error)."""
    ws.title = "Per-image detail"
    headers = ["No.", "Image source", "Ground-truth style",
               "System prediction", "Top-1 result", "Panel agreement",
               "Abstained?"]
    _header_row(ws, 1, headers)

    def _gen_label(gen: str) -> str:
        return "ChatGPT" if gen.strip().lower() == "chatgpt" else "Gemini"

    def _sort_key(r: dict) -> Tuple[int, str]:
        try:
            pid = int(r["prompt_id"])
        except (KeyError, ValueError):
            pid = 0
        return pid, r.get("generator", "")

    row = 2
    for r in sorted(m.detail, key=_sort_key):
        is_error = bool(r.get("error")) or int(r.get("n_candidates") or 0) <= 0
        agree = r.get("panel_agreement")
        agree_val = round(float(agree), 3) if agree not in (None, "", "None") else ""
        if is_error:
            result_txt = "API ERROR"
        else:
            result_txt = "Correct" if int(r["top1_hit"]) else "Wrong"
        abstain_txt = "Yes" if (not is_error and int(r.get("uncertain") or 0)) else ""

        values = [
            int(r["prompt_id"]) if r.get("prompt_id", "").isdigit() else r.get("prompt_id"),
            _gen_label(r.get("generator", "")),
            r.get("gt_style", ""),
            r.get("pred_style", ""),
            result_txt,
            agree_val,
            abstain_txt,
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _LEFT if col in (3, 4) else _CENTER
        # Colour the result cell.
        rc = ws.cell(row=row, column=5)
        if is_error:
            rc.fill = _ERR_FILL
        elif result_txt == "Correct":
            rc.fill, rc.font = _OK_FILL, _OK_FONT
        else:
            rc.fill, rc.font = _BAD_FILL, _BAD_FONT
        row += 1

    _autofit(ws, {1: 7, 2: 11, 3: 32, 4: 32, 5: 14, 6: 14, 7: 10})
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def _write_guide_sheet(ws: Worksheet, blocks: List[tuple]) -> None:
    """Write a column dictionary: for each sheet, every column + what it means.

    ``blocks`` is a list of ``(section_title, [(column, meaning, cell_values)])``.
    """
    ws.title = "Column guide"
    ws["A1"] = "COLUMN GUIDE - meaning of every column and its cell values"
    ws["A1"].font = _TITLE_FONT
    row = 3
    for section, cols in blocks:
        ws.cell(row=row, column=1, value=section).font = Font(bold=True, size=12)
        row += 1
        _header_row(ws, row, ["Column", "Meaning", "Cell values"])
        row += 1
        for col, meaning, values in cols:
            for c, v in enumerate([col, meaning, values], start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = _BORDER
                cell.alignment = _LEFT
            row += 1
        row += 1
    _autofit(ws, {1: 22, 2: 62, 3: 40})
    ws.sheet_view.showGridLines = False


_RUN_GUIDE_BLOCKS = [
    ("Sheet 'Summary' - top-1 accuracy table", [
        ("Group", "Which subset of images the row summarises.",
         "OVERALL (all 200) / ChatGPT images (100) / Gemini images (100)"),
        ("Correct", "Number of images whose top-1 prediction equals the "
         "ground-truth style.", "integer count"),
        ("Scored", "Images actually analysed; API/quota-failed images are "
         "excluded.", "integer count"),
        ("Top-1 (%)", "Correct / Scored x 100.", "percentage 0-100"),
    ]),
    ("Sheet 'Summary' - risk-coverage table", [
        ("Threshold t", "Abstention cutoff on panel agreement: results with "
         "agreement below t are refused.", "0.0 - 0.8"),
        ("Kept", "Images kept (panel agreement >= t, not refused).",
         "integer count"),
        ("Coverage (%)", "Kept / Scored x 100 - how many images still get an "
         "answer.", "percentage 0-100"),
        ("Selective Top-1 (%)", "Top-1 accuracy computed on the kept images "
         "only.", "percentage 0-100"),
    ]),
    ("Sheet 'Per-image detail' - one row per image", [
        ("No.", "Benchmark image id (prompt_id).", "1 - 100"),
        ("Image source", "Which model generated the image.",
         "ChatGPT / Gemini"),
        ("Ground-truth style", "Correct architectural style (from "
         "ground_truth.csv).", "style name"),
        ("System prediction", "Style the system returned at rank 1.",
         "style name (empty if API error)"),
        ("Top-1 result", "Whether prediction matches the ground truth.",
         "Correct (green) / Wrong (red) / API ERROR (grey)"),
        ("Panel agreement", "Average pairwise Spearman rank correlation across "
         "the 3 vision judges; higher = stronger agreement.",
         "0.0 - 1.0 (empty if API error)"),
        ("Abstained?", "System flagged the result as uncertain "
         "(low agreement / margin).", "Yes / empty"),
    ]),
]

_AGG_GUIDE_BLOCKS = [
    ("Sheet 'Summary across runs'", [
        ("Run", "One full 200-image evaluation pass. Runs differ because the "
         "language models are stochastic.", "run label + date"),
        ("Top-1 OVERALL (%)", "Top-1 accuracy over all 200 images.",
         "percentage 0-100"),
        ("Top-1 ChatGPT (%)", "Top-1 accuracy over the 100 ChatGPT images.",
         "percentage 0-100"),
        ("Top-1 Gemini (%)", "Top-1 accuracy over the 100 Gemini images.",
         "percentage 0-100"),
        ("API errors", "Images excluded from scoring due to API/quota failure.",
         "integer count"),
        ("Mean +/- Standard deviation", "Mean and sample standard deviation of "
         "each column across the runs.", "mean +/- sd"),
    ]),
]


def _export_run(csv_path: Path, xlsx_path: Path, label: str) -> None:
    """Build a workbook for a single run (summary + detail + column guide)."""
    m = _load_metrics(csv_path)
    wb = openpyxl.Workbook()
    _write_summary_sheet(wb.active, m, label)
    _write_detail_sheet(wb.create_sheet(), m)
    _write_guide_sheet(wb.create_sheet(), _RUN_GUIDE_BLOCKS)
    _save(wb, xlsx_path)
    print(f"[run] {label}: top-1 {_pct_str(m.top1, m.n_scored)}"
          f"  (errors: {m.n_errors})  -> {xlsx_path}")


# ---- aggregate across runs -------------------------------------------------
def _export_aggregate(csv_paths: List[Path], labels: List[str],
                      xlsx_path: Path) -> None:
    """Build a workbook comparing several runs with mean +/- std."""
    mets = [_load_metrics(p) for p in csv_paths]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary across runs"
    ws["A1"] = "EXPERIMENT SUMMARY ACROSS RUNS (Top-1)"
    ws["A1"].font = _TITLE_FONT
    ws["A2"] = ("The pipeline is stochastic (LLM temperature > 0); the standard "
                "deviation reflects run-to-run variation.")
    ws["A2"].font = _SUB_FONT

    _header_row(ws, 4, ["Run", "Top-1 OVERALL (%)", "Top-1 ChatGPT (%)",
                        "Top-1 Gemini (%)", "API errors"])
    overall, chat, gem = [], [], []
    row = 5
    for label, m in zip(labels, mets):
        o = m.pct(m.top1)
        c = m.pct(m.by_gen.get("chatgpt", {}).get("top1", 0),
                  m.by_gen.get("chatgpt", {}).get("n", 0))
        g = m.pct(m.by_gen.get("gemini", {}).get("top1", 0),
                  m.by_gen.get("gemini", {}).get("n", 0))
        overall.append(o); chat.append(c); gem.append(g)
        for col, v in enumerate([label, round(o, 1), round(c, 1), round(g, 1),
                                 m.n_errors], start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _LEFT if col == 1 else _CENTER
        row += 1

    def _mean_std(xs: List[float]) -> str:
        if not xs:
            return "n/a"
        sd = statistics.stdev(xs) if len(xs) > 1 else 0.0
        return f"{statistics.mean(xs):.1f} ± {sd:.1f}"

    summary = ["Mean ± Standard deviation",
               _mean_std(overall), _mean_std(chat), _mean_std(gem), ""]
    for col, v in enumerate(summary, start=1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER
        cell.alignment = _LEFT if col == 1 else _CENTER

    # Risk-coverage averaged over runs.
    row += 2
    ws.cell(row=row, column=1,
            value="Risk-coverage averaged across runs").font = Font(bold=True, size=12)
    row += 1
    _header_row(ws, row, ["Threshold t", "Mean coverage (%)",
                          "Mean selective Top-1 (%)"])
    row += 1
    for i, t in enumerate(_RISK_THRESHOLDS):
        covs, sels = [], []
        for m in mets:
            _, kept, sel = m.risk_rows[i]
            if m.n_scored:
                covs.append(100.0 * kept / m.n_scored)
            if kept:
                sels.append(100.0 * sel / kept)
        cov = round(statistics.mean(covs), 1) if covs else 0.0
        sel = round(statistics.mean(sels), 1) if sels else 0.0
        for col, v in enumerate([t, cov, sel], start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _CENTER
        row += 1

    _autofit(ws, {1: 30, 2: 18, 3: 20, 4: 18, 5: 16})
    ws.sheet_view.showGridLines = False
    _write_guide_sheet(wb.create_sheet(), _AGG_GUIDE_BLOCKS)
    _save(wb, xlsx_path)
    print(f"[aggregate] {len(mets)} runs -> {xlsx_path}\n"
          f"  Top-1 OVERALL : {_mean_std(overall)}\n"
          f"  Top-1 ChatGPT : {_mean_std(chat)}\n"
          f"  Top-1 Gemini  : {_mean_std(gem)}")


def _save(wb: openpyxl.Workbook, xlsx_path: Path) -> None:
    """Save the workbook, giving a clear message if the file is open in Excel."""
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(xlsx_path)
    except PermissionError:
        print(f"ERROR: cannot write {xlsx_path} - close it in Excel first.",
              file=sys.stderr)
        raise


def main() -> int:
    """Parse CLI args and export the requested workbook(s)."""
    parser = argparse.ArgumentParser(description="Export eval results to .xlsx")
    parser.add_argument("--mode", required=True, choices=["run", "aggregate"])
    parser.add_argument("--csv", help="[run] single eval CSV")
    parser.add_argument("--xlsx", required=True, help="output .xlsx path")
    parser.add_argument("--label", default="", help="[run] run label")
    parser.add_argument("--csvs", help="[aggregate] comma-separated eval CSVs")
    parser.add_argument("--labels", help="[aggregate] ';'-separated run labels")
    args = parser.parse_args()

    if args.mode == "run":
        if not args.csv:
            parser.error("--csv is required for --mode run")
        _export_run(Path(args.csv), Path(args.xlsx),
                    args.label or Path(args.csv).stem)
    else:
        if not args.csvs:
            parser.error("--csvs is required for --mode aggregate")
        paths = [Path(p.strip()) for p in args.csvs.split(",") if p.strip()]
        labels = ([s.strip() for s in args.labels.split(";")]
                  if args.labels else [p.stem for p in paths])
        if len(labels) != len(paths):
            parser.error("number of --labels must match number of --csvs")
        _export_aggregate(paths, labels, Path(args.xlsx))
    return 0


if __name__ == "__main__":
    sys.exit(main())
